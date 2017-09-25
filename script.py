import openpyxl

dest_filename = 'UserList.xlsx'

wb = openpyxl.load_workbook(dest_filename)

#worksheets = wb.get_sheet_names()
#ws = worksheets[0]
#print(ws)

ws= wb.worksheets[0]

#print(str(ws.max_row)  + "  " +  str(ws.max_column))

# Solution 1:
for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        print(cell.value, end=" ")
    print()    